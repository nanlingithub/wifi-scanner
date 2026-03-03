#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
企业级报告生成标签页
版本: 1.7 - 集成缓存/日志/配置系统
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime
import os
import threading
import subprocess
import time

from wifi_modules.theme import ModernTheme, ModernButton, ModernCard
from wifi_modules.enterprise_signal_analyzer import EnterpriseSignalAnalyzer
from wifi_modules.pci_dss_security_assessor import PCIDSSSecurityAssessor
from wifi_modules.enterprise_reports import PDFGeneratorAsync
from wifi_modules.enterprise_reports.templates.signal_template import SignalAnalysisTemplate
from wifi_modules.enterprise_reports.templates.security_template import SecurityAssessmentTemplate
from wifi_modules.enterprise_reports.templates.pci_dss_template import PCIDSSTemplate

# 新增：导入缓存、日志和配置系统
from wifi_modules.cache_manager import NetworkAnalysisCache
from wifi_modules.logger import get_logger
from wifi_modules.config_loader import ConfigLoader


class EnterpriseReportTab:
    """企业级报告生成标签页"""
    
    def __init__(self, parent, wifi_analyzer):
        self.parent = parent
        self.wifi_analyzer = wifi_analyzer
        self.frame = ttk.Frame(parent)
        
        # 初始化分析器
        self.signal_analyzer = EnterpriseSignalAnalyzer()
        self.security_assessor = PCIDSSSecurityAssessor()
        # 使用 v2.0 PDF生成器（统一接口 + 智能缓存）
        self.report_generator = PDFGeneratorAsync(use_cache=True)
        self.signal_template = SignalAnalysisTemplate(self.report_generator.styles)
        self.security_template = SecurityAssessmentTemplate(self.report_generator.styles)
        self.pci_template = PCIDSSTemplate(self.report_generator.styles)
        
        # 新增：初始化缓存系统
        self.cache = NetworkAnalysisCache()
        
        # 新增：初始化日志系统
        self.logger = get_logger('EnterpriseReportTab')
        self.logger.info("企业级报告标签页初始化")
        
        # 新增：加载配置
        self.config = ConfigLoader()
        
        self.current_analysis = None
        
        # 新增：多点位采集数据存储
        self.collection_points = []  # 格式: [{'location': str, 'timestamp': str, 'networks': list}, ...]
        self.logger.info("多点位采集系统已初始化")
        self.current_assessment = None
        
        # 持久化配置文件路径
        self.config_dir = os.path.join(os.path.dirname(__file__), '..', 'config')
        self.auth_ssid_file = os.path.join(self.config_dir, 'authorized_ssids.json')
        self.multipoint_file = os.path.join(self.config_dir, 'multipoint_collections.json')
        
        # 确保配置目录存在
        os.makedirs(self.config_dir, exist_ok=True)
        
        # 加载保存的配置
        self.saved_auth_ssids = self._load_auth_ssids()
        self._load_multipoint_data()
        
        self._setup_ui()
    
    def _setup_ui(self):
        """设置UI（优化版 - 添加滚动条）"""
        # 主容器 - 添加渐变背景
        main_container = tk.Frame(self.frame, bg='#f0f4f8')
        main_container.pack(fill='both', expand=True)
        
        # 顶部标题栏 - 优化高度和视觉层次
        title_bar = tk.Frame(main_container, bg='#2c3e50', height=70)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)
        
        # 主标题 - 更大更醒目
        title_label = tk.Label(
            title_bar,
            text="📊 企业级网络分析与安全评估中心",
            font=('Microsoft YaHei UI', 18, 'bold'),
            bg='#2c3e50',
            fg='white'
        )
        title_label.pack(pady=(12, 2))
        
        # 副标题 - 增大字体
        subtitle_label = tk.Label(
            title_bar,
            text="Enterprise Network Analysis & Security Assessment Platform",
            font=('Arial', 9),
            bg='#2c3e50',
            fg='#bdc3c7'
        )
        subtitle_label.pack()
        
        # 创建滚动区域容器 - 移除边距
        scroll_container = tk.Frame(main_container, bg='#f0f4f8')
        scroll_container.pack(fill='both', expand=True, padx=0, pady=0)
        
        # 创建Canvas和Scrollbar
        canvas = tk.Canvas(scroll_container, bg='#f0f4f8', highlightthickness=0)
        scrollbar = ttk.Scrollbar(scroll_container, orient='vertical', command=canvas.yview)
        
        # 创建可滚动的内容框架
        scrollable_frame = tk.Frame(canvas, bg='#f0f4f8')
        
        # 绑定滚动事件
        scrollable_frame.bind(
            '<Configure>',
            lambda e: canvas.configure(scrollregion=canvas.bbox('all'))
        )
        
        # 在Canvas中创建窗口
        canvas_window = canvas.create_window((0, 0), window=scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)
        
        # 绑定Canvas宽度变化事件，确保内容填充整个宽度和高度
        def _configure_canvas_width(event):
            # 设置scrollable_frame的宽度与canvas一致
            canvas.itemconfig(canvas_window, width=event.width)
            # 同时设置高度，确保内容至少占满整个canvas高度
            canvas_height = event.height
            content_height = scrollable_frame.winfo_reqheight()
            # 使用两者中的较大值
            canvas.itemconfig(canvas_window, height=max(canvas_height, content_height))
        canvas.bind('<Configure>', _configure_canvas_width)
        
        # 布局Canvas和Scrollbar
        canvas.pack(side='left', fill='both', expand=True)
        scrollbar.pack(side='right', fill='y')
        
        # 鼠标滚轮支持
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        canvas.bind_all("<MouseWheel>", _on_mousewheel)
        
        # 内容区域（使用scrollable_frame而不是main_container）- 完全移除边距
        content_frame = tk.Frame(scrollable_frame, bg='#f0f4f8')
        content_frame.pack(fill='both', expand=True, padx=0, pady=0)
        
        # 创建三列布局 - 使用Grid确保比例控制
        columns_frame = tk.Frame(content_frame, bg='#f0f4f8')
        columns_frame.pack(fill='both', expand=True)
        
        # 配置Grid列权重（信号:安全:操作 = 1:1:1，完全等宽）
        columns_frame.grid_columnconfigure(0, weight=1, uniform='equal')
        columns_frame.grid_columnconfigure(1, weight=1, uniform='equal')
        columns_frame.grid_columnconfigure(2, weight=1, uniform='equal')
        columns_frame.grid_rowconfigure(0, weight=1)  # 行权重确保垂直填充
        
        # 左列 - 信号分析报告（移除所有边距）
        left_column = tk.Frame(columns_frame, bg='#f0f4f8')
        left_column.grid(row=0, column=0, sticky='nsew', padx=0)
        
        self._create_signal_analysis_section(left_column)
        
        # 中列 - 安全评估报告
        middle_column = tk.Frame(columns_frame, bg='#f0f4f8')
        middle_column.grid(row=0, column=1, sticky='nsew', padx=0)
        
        self._create_security_assessment_section(middle_column)
        
        # 右列 - 综合报告操作中心（移除所有边距）
        right_column = tk.Frame(columns_frame, bg='#f0f4f8')
        right_column.grid(row=0, column=2, sticky='nsew', padx=0)
        
        self._create_actions_section(right_column)
    
    def _create_signal_analysis_section(self, parent):
        """创建信号分析报告区域（优化版）"""
        # 卡片外层容器 - 增强阴影效果
        shadow_frame = tk.Frame(parent, bg='#c8d6e5')
        shadow_frame.pack(fill='both', expand=True, padx=2, pady=2)
        
        # 卡片容器 - 白色背景
        card_frame = tk.Frame(shadow_frame, bg='white', relief='flat', borderwidth=0)
        card_frame.pack(fill='both', expand=True)
        
        # 卡片标题栏 - 优化视觉比例
        title_bar = tk.Frame(card_frame, bg='#3498db', height=50)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)
        
        title_icon_label = tk.Label(
            title_bar,
            text="🌐",
            font=('Segoe UI Emoji', 22),
            bg='#3498db',
            fg='white'
        )
        title_icon_label.pack(side='left', padx=15)
        
        title_text_frame = tk.Frame(title_bar, bg='#3498db')
        title_text_frame.pack(side='left', fill='y', pady=8)
        
        tk.Label(
            title_text_frame,
            text="WiFi信号质量分析",
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg='#3498db',
            fg='white'
        ).pack(anchor='w')
        
        tk.Label(
            title_text_frame,
            text="Signal Quality Analysis Report",
            font=('Arial', 8),
            bg='#3498db',
            fg='#ecf0f1'
        ).pack(anchor='w')
        
        # 内容区域 - 优化布局
        content_area = tk.Frame(card_frame, bg='white')
        content_area.pack(fill='both', expand=True, padx=12, pady=8)
        
        # 功能说明区 - 横向紧凑布局（节省空间）
        features_frame = tk.Frame(content_area, bg='#e3f2fd', relief='flat')
        features_frame.pack(fill='x', pady=(0, 8))
        
        features_inner = tk.Frame(features_frame, bg='#e3f2fd')
        features_inner.pack(pady=6, padx=8)
        
        features = [
            ("📊", "信号评估"),
            ("📡", "覆盖分析"),
            ("📶", "信道监控"),
            ("⚡", "干扰检测")
        ]
        
        for i, (icon, title) in enumerate(features):
            item_label = tk.Label(
                features_inner,
                text=f"{icon}{title}",
                font=('Microsoft YaHei UI', 8),
                bg='#e3f2fd',
                fg='#1976d2',
                padx=8
            )
            item_label.pack(side='left', padx=3)
        
        # 功能介绍卡片 - 增加详细信息
        intro_frame = tk.Frame(content_area, bg='#f0f8ff', relief='flat', borderwidth=1)
        intro_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(
            intro_frame,
            text="📝 分析说明：基于IEEE 802.11标准，对WiFi信号质量、覆盖范围、\n   信道利用率和干扰源进行全面检测。提供专业的性能评估报告。",
            font=('Microsoft YaHei UI', 9),
            bg='#f0f8ff',
            fg='#1565c0',
            justify='left',
            pady=5,
            padx=8
        ).pack(anchor='w')
        
        # 多点位采集管理区域 - 企业级新功能
        multipoint_container = tk.Frame(content_area, bg='#fff3e0', relief='solid', borderwidth=1)
        multipoint_container.pack(fill='x', pady=(0, 8))
        
        mp_header = tk.Frame(multipoint_container, bg='#fff3e0')
        mp_header.pack(fill='x', padx=10, pady=(6, 3))
        
        tk.Label(
            mp_header,
            text="📍",
            font=('Segoe UI Emoji', 12),
            bg='#fff3e0'
        ).pack(side='left', padx=(0, 5))
        
        tk.Label(
            mp_header,
            text="多点位信号采集:",
            font=('Microsoft YaHei UI', 10, 'bold'),
            bg='#fff3e0',
            fg='#e65100'
        ).pack(side='left')
        
        mp_input_frame = tk.Frame(multipoint_container, bg='#fff3e0')
        mp_input_frame.pack(fill='x', padx=10, pady=(0, 6))
        
        tk.Label(
            mp_input_frame,
            text="位置名称:",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#fff3e0',
            fg='#f57c00'
        ).pack(side='left', padx=(0, 5))
        
        self.location_entry = tk.Entry(
            mp_input_frame,
            width=20,
            font=('Microsoft YaHei UI', 9),
            relief='solid',
            borderwidth=1
        )
        self.location_entry.pack(side='left', padx=(0, 8))
        self.location_entry.insert(0, "办公区A-1层")
        
        add_point_btn = tk.Button(
            mp_input_frame,
            text="📊 采集此点",
            command=self._add_collection_point,
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#ff9800',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=12,
            pady=4
        )
        add_point_btn.pack(side='left', padx=(0, 8))
        
        self.points_label = tk.Label(
            mp_input_frame,
            text=f"已采集: {len(self.collection_points)} 点",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#fff3e0',
            fg='#e65100'
        )
        self.points_label.pack(side='left')
        
        # 多点位数据管理按钮区域
        mp_btn_frame = tk.Frame(multipoint_container, bg='#fff3e0')
        mp_btn_frame.pack(fill='x', padx=10, pady=(0, 6))
        
        save_mp_btn = tk.Button(
            mp_btn_frame,
            text="💾 保存采集",
            command=self._save_multipoint_data,
            font=('Microsoft YaHei UI', 9),
            bg='#ff9800',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=10,
            pady=4
        )
        save_mp_btn.pack(side='left', padx=(0, 5))
        
        load_mp_btn = tk.Button(
            mp_btn_frame,
            text="📂 加载采集",
            command=self._load_multipoint_data_from_file,
            font=('Microsoft YaHei UI', 9),
            bg='#42a5f5',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=10,
            pady=4
        )
        load_mp_btn.pack(side='left', padx=(0, 5))
        
        clear_mp_btn = tk.Button(
            mp_btn_frame,
            text="🗑️ 清空",
            command=self._clear_multipoint_data,
            font=('Microsoft YaHei UI', 9),
            bg='#ef5350',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=10,
            pady=4
        )
        clear_mp_btn.pack(side='left', padx=(0, 5))
        
        chart_btn = tk.Button(
            mp_btn_frame,
            text="📈 对比图表",
            command=self._generate_multipoint_chart,
            font=('Microsoft YaHei UI', 9),
            bg='#9c27b0',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=10,
            pady=4
        )
        chart_btn.pack(side='left')
        
        # 按钮区域 - 统一布局
        button_area = tk.Frame(content_area, bg='white')
        button_area.pack(fill='both', expand=True, pady=(10, 0))
        
        # 主操作按钮 - 视觉焦点（统一高度：16px）
        analyze_btn = tk.Button(
            button_area,
            text="🔍  执行信号分析",
            command=self._run_signal_analysis,
            font=('Microsoft YaHei UI', 11, 'bold'),
            bg='#3498db',
            fg='white',
            activebackground='#2980b9',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=16,
            borderwidth=0,
            highlightthickness=0
        )
        analyze_btn.pack(fill='x', pady=(0, 10), padx=2)
        
        # 鼠标悬停效果 - 微妙反馈
        def on_enter_analyze(e):
            analyze_btn.config(bg='#2980b9')
        def on_leave_analyze(e):
            analyze_btn.config(bg='#3498db')
        def on_click_analyze(e):
            analyze_btn.config(bg='#1f618d')
        
        analyze_btn.bind("<Enter>", on_enter_analyze)
        analyze_btn.bind("<Leave>", on_leave_analyze)
        analyze_btn.bind("<Button-1>", on_click_analyze)
        
        # 进度指示区 - 详细进度展示
        self.signal_progress_frame = tk.Frame(content_area, bg='white')
        self.signal_progress_frame.pack(fill='x', pady=(8, 5))
        
        # 进度条容器
        progress_container = tk.Frame(self.signal_progress_frame, bg='#ecf0f1', relief='solid', borderwidth=1)
        progress_container.pack(fill='x', pady=2)
        
        self.signal_progress = ttk.Progressbar(
            progress_container,
            mode='determinate',
            length=400,
            maximum=100
        )
        self.signal_progress.pack(fill='x', padx=5, pady=5)
        
        # 进度百分比和阶段提示
        progress_info_frame = tk.Frame(progress_container, bg='#ecf0f1')
        progress_info_frame.pack(fill='x', padx=5, pady=(0, 5))
        
        self.signal_progress_percent = tk.Label(
            progress_info_frame,
            text="0%",
            font=('Microsoft YaHei UI', 8, 'bold'),
            bg='#ecf0f1',
            fg='#3498db'
        )
        self.signal_progress_percent.pack(side='left')
        
        self.signal_progress_label = tk.Label(
            progress_info_frame,
            text="",
            font=('Microsoft YaHei UI', 8),
            bg='#ecf0f1',
            fg='#7f8c8d'
        )
        self.signal_progress_label.pack(side='left', padx=10)
        
        # 初始隐藏进度区域
        self.signal_progress_frame.pack_forget()
        
        # 次要操作按钮 - 适中尺寸
        self.signal_preview_btn = tk.Button(
            button_area,
            text="👁️  预览分析结果",
            command=self._preview_signal_analysis,
            font=('Microsoft YaHei UI', 10),
            bg='#27ae60',
            fg='white',
            activebackground='#229954',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=14,
            borderwidth=0,
            highlightthickness=0,
            state='disabled'
        )
        self.signal_preview_btn.pack(fill='x', pady=0, padx=0)
        
        # 悬停效果
        def on_enter_preview(e):
            if self.signal_preview_btn['state'] == 'normal':
                self.signal_preview_btn.config(bg='#229954')
        def on_leave_preview(e):
            if self.signal_preview_btn['state'] == 'normal':
                self.signal_preview_btn.config(bg='#27ae60')
        self.signal_preview_btn.bind("<Enter>", on_enter_preview)
        self.signal_preview_btn.bind("<Leave>", on_leave_preview)
        
        # 快速统计信息卡片 - 超紧凑设计
        self.signal_stats_frame = tk.Frame(content_area, bg='white')
        self.signal_stats_frame.pack(fill='x', pady=3)
        
        stats_container = tk.Frame(self.signal_stats_frame, bg='#ecf0f1', relief='flat')
        stats_container.pack(fill='x')
        
        # 三列统计布局 - 超紧凑卡片
        col1 = tk.Frame(stats_container, bg='#e8f5e9', relief='solid', borderwidth=1)
        col1.pack(side='left', fill='both', expand=True, padx=2, pady=3)
        
        tk.Label(
            col1,
            text="📈",
            font=('Segoe UI Emoji', 10),
            bg='#e8f5e9'
        ).pack(pady=(2, 0))
        
        self.signal_score_label = tk.Label(
            col1,
            text="评分\n--",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#e8f5e9',
            fg='#27ae60'
        )
        self.signal_score_label.pack(pady=(0, 2))
        
        col2 = tk.Frame(stats_container, bg='#e3f2fd', relief='solid', borderwidth=1)
        col2.pack(side='left', fill='both', expand=True, padx=2, pady=3)
        
        tk.Label(
            col2,
            text="📊",
            font=('Segoe UI Emoji', 10),
            bg='#e3f2fd'
        ).pack(pady=(2, 0))
        
        self.signal_avg_label = tk.Label(
            col2,
            text="信号\n--",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#e3f2fd',
            fg='#2980b9'
        )
        self.signal_avg_label.pack(pady=(0, 2))
        
        col3 = tk.Frame(stats_container, bg='#fff3e0', relief='solid', borderwidth=1)
        col3.pack(side='left', fill='both', expand=True, padx=2, pady=3)
        
        tk.Label(
            col3,
            text="⚡",
            font=('Segoe UI Emoji', 10),
            bg='#fff3e0'
        ).pack(pady=(2, 0))
        
        self.signal_stability_label = tk.Label(
            col3,
            text="稳定\n--",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#fff3e0',
            fg='#f39c12'
        )
        self.signal_stability_label.pack(pady=(0, 2))
        
        # 初始隐藏统计信息
        self.signal_stats_frame.pack_forget()
        
        # 导出按钮区 - 统一配色（橙色PDF + 绿色Excel + 蓝色JSON）
        export_frame = tk.Frame(content_area, bg='white')
        export_frame.pack(fill='x', pady=(0, 10))
        
        # PDF导出按钮 - 统一橙色
        pdf_btn = tk.Button(
            export_frame,
            text="📄 PDF",
            command=self._export_signal_pdf,
            font=('Microsoft YaHei UI', 8, 'bold'),
            bg='#e67e22',
            fg='white',
            activebackground='#d35400',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=8,
            borderwidth=0
        )
        pdf_btn.pack(side='left', fill='x', expand=True, padx=(2, 2))
        
        # Excel导出按钮 - 绿色
        excel_btn = tk.Button(
            export_frame,
            text="📊 Excel",
            command=self._export_signal_excel,
            font=('Microsoft YaHei UI', 8, 'bold'),
            bg='#27ae60',
            fg='white',
            activebackground='#229954',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=8,
            borderwidth=0
        )
        excel_btn.pack(side='left', fill='x', expand=True, padx=(2, 2))
        
        # JSON导出按钮 - 统一蓝色
        json_btn = tk.Button(
            export_frame,
            text="💾 JSON",
            command=self._export_signal_json,
            font=('Microsoft YaHei UI', 8, 'bold'),
            bg='#3498db',
            fg='white',
            activebackground='#2980b9',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=8,
            borderwidth=0
        )
        json_btn.pack(side='left', fill='x', expand=True, padx=(2, 2))
        
        # 悬停效果
        def on_enter_pdf(e): pdf_btn.config(bg='#d35400')
        def on_leave_pdf(e): pdf_btn.config(bg='#e67e22')
        def on_enter_excel(e): excel_btn.config(bg='#229954')
        def on_leave_excel(e): excel_btn.config(bg='#27ae60')
        def on_enter_json(e): json_btn.config(bg='#2980b9')
        def on_leave_json(e): json_btn.config(bg='#3498db')
        
        pdf_btn.bind("<Enter>", on_enter_pdf)
        pdf_btn.bind("<Leave>", on_leave_pdf)
        excel_btn.bind("<Enter>", on_enter_excel)
        excel_btn.bind("<Leave>", on_leave_excel)
        json_btn.bind("<Enter>", on_enter_json)
        json_btn.bind("<Leave>", on_leave_json)
        
        # 状态标签 - 清晰反馈
        status_container = tk.Frame(content_area, bg='#f8f9fa', relief='ridge', borderwidth=2)
        status_container.pack(fill='x', pady=(0, 10))
        
        self.signal_status_label = tk.Label(
            status_container,
            text="💤 等待分析",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#f8f9fa',
            fg='#7f8c8d',
            pady=8
        )
        self.signal_status_label.pack()
    
    def _create_security_assessment_section(self, parent):
        """创建安全评估报告区域（美化版）"""
        # 卡片外层容器 - 实现阴影效果
        shadow_frame = tk.Frame(parent, bg='#d0d8e0')
        shadow_frame.pack(fill='both', expand=True, padx=2, pady=2)
        
        # 卡片容器 - 白色背景
        card_frame = tk.Frame(shadow_frame, bg='white', relief='flat', borderwidth=0)
        card_frame.pack(fill='both', expand=True)
        
        # 卡片标题栏 - 红色警告性主题
        title_bar = tk.Frame(card_frame, bg='#e74c3c', height=50)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)
        
        title_icon_label = tk.Label(
            title_bar,
            text="🔒",
            font=('Segoe UI Emoji', 22),
            bg='#e74c3c',
            fg='white'
        )
        title_icon_label.pack(side='left', padx=15)
        
        title_text_frame = tk.Frame(title_bar, bg='#e74c3c')
        title_text_frame.pack(side='left', fill='y', pady=8)
        
        tk.Label(
            title_text_frame,
            text="PCI-DSS安全风险评估",
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg='#e74c3c',
            fg='white'
        ).pack(anchor='w')
        
        tk.Label(
            title_text_frame,
            text="Wireless Security Risk Assessment",
            font=('Arial', 8),
            bg='#e74c3c',
            fg='#fce4ec'
        ).pack(anchor='w')
        
        # 内容区域 - 优化布局
        content_area = tk.Frame(card_frame, bg='white')
        content_area.pack(fill='both', expand=True, padx=12, pady=8)
        
        # 功能说明区 - 横向紧凑布局
        features_frame = tk.Frame(content_area, bg='#ffebee', relief='flat')
        features_frame.pack(fill='x', pady=(0, 8))
        
        features_inner = tk.Frame(features_frame, bg='#ffebee')
        features_inner.pack(pady=6, padx=8)
        
        features = [
            ("🔐", "加密检测"),
            ("🚨", "未授权AP"),
            ("⚠️", "配置审查"),
            ("✅", "合规性")
        ]
        
        for icon, title in features:
            item_label = tk.Label(
                features_inner,
                text=f"{icon}{title}",
                font=('Microsoft YaHei UI', 8),
                bg='#ffebee',
                fg='#c62828',
                padx=6
            )
            item_label.pack(side='left', padx=3)
        
        # 功能介绍卡片 - 增加详细信息
        intro_frame = tk.Frame(content_area, bg='#fff5f5', relief='flat', borderwidth=1)
        intro_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(
            intro_frame,
            text="📝 评估说明：依据PCI-DSS安全标准，检测加密协议强度、\n   识别非授权接入点、审查配置安全性，生成风险评估报告。",
            font=('Microsoft YaHei UI', 9),
            bg='#fff5f5',
            fg='#c62828',
            justify='left',
            pady=5,
            padx=8
        ).pack(anchor='w')
        
        # 授权SSID配置区域 - 企业级管理（优化：从AP改为SSID）
        auth_container = tk.Frame(content_area, bg='#e8f5e9', relief='solid', borderwidth=1)
        auth_container.pack(fill='x', pady=5)
        
        auth_inner = tk.Frame(auth_container, bg='#e8f5e9')
        auth_inner.pack(fill='x', padx=10, pady=6)
        
        tk.Label(
            auth_inner,
            text="✅",
            font=('Segoe UI Emoji', 12),
            bg='#e8f5e9'
        ).pack(side='left', padx=(0, 5))
        
        tk.Label(
            auth_inner,
            text="授权WiFi SSID:",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#e8f5e9',
            fg='#2e7d32'
        ).pack(side='left')
        
        self.auth_ssid_entry = tk.Entry(
            auth_inner,
            width=28,
            font=('Microsoft YaHei UI', 9),
            relief='solid',
            borderwidth=1
        )
        self.auth_ssid_entry.pack(side='left', padx=5, fill='x', expand=True)
        
        # 加载保存的授权SSID
        if self.saved_auth_ssids:
            self.auth_ssid_entry.insert(0, ', '.join(self.saved_auth_ssids))
        
        # 保存按钮
        save_ssid_btn = tk.Button(
            auth_inner,
            text="💾",
            command=self._save_auth_ssids,
            font=('Segoe UI Emoji', 10),
            bg='#66bb6a',
            fg='white',
            cursor='hand2',
            relief='flat',
            padx=5,
            pady=1
        )
        save_ssid_btn.pack(side='left', padx=(3, 0))
        
        tk.Label(
            auth_inner,
            text="SSID用逗号分隔",
            font=('Microsoft YaHei UI', 8),
            bg='#e8f5e9',
            fg='#66bb6a'
        ).pack(side='left')
        
        # 按钮区域 - 统一布局
        button_area = tk.Frame(content_area, bg='white')
        button_area.pack(fill='both', expand=True, pady=(10, 0))
        
        # 主评估按钮 - 视觉焦点（统一高度：16px）
        assess_btn = tk.Button(
            button_area,
            text="🛡️  执行安全评估",
            command=self._run_security_assessment,
            font=('Microsoft YaHei UI', 11, 'bold'),
            bg='#e74c3c',
            fg='white',
            activebackground='#c0392b',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=16,
            borderwidth=0,
            highlightthickness=0
        )
        assess_btn.pack(fill='x', pady=(0, 10), padx=2)
        
        # 悬停效果 - 微妙反馈
        def on_enter_assess(e): 
            assess_btn.config(bg='#c0392b')
        def on_leave_assess(e): 
            assess_btn.config(bg='#e74c3c')
        def on_click_assess(e):
            assess_btn.config(bg='#a93226')
        assess_btn.bind("<Enter>", on_enter_assess)
        assess_btn.bind("<Leave>", on_leave_assess)
        assess_btn.bind("<Button-1>", on_click_assess)
        
        # 次要操作按钮（统一高度：12px）
        self.security_preview_btn = tk.Button(
            button_area,
            text="👁️  预览评估结果",
            command=self._preview_security_assessment,
            font=('Microsoft YaHei UI', 10),
            bg='#27ae60',
            fg='white',
            activebackground='#229954',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=12,
            borderwidth=0,
            highlightthickness=0,
            state='disabled'
        )
        self.security_preview_btn.pack(fill='x', pady=(0, 10), padx=2)
        
        # 悬停效果
        def on_enter_sec_preview(e):
            if self.security_preview_btn['state'] == 'normal':
                self.security_preview_btn.config(bg='#229954')
        def on_leave_sec_preview(e):
            if self.security_preview_btn['state'] == 'normal':
                self.security_preview_btn.config(bg='#27ae60')
        self.security_preview_btn.bind("<Enter>", on_enter_sec_preview)
        self.security_preview_btn.bind("<Leave>", on_leave_sec_preview)
        
        # 导出按钮区 - 统一高度（8px）
        export_frame = tk.Frame(content_area, bg='white')
        export_frame.pack(fill='x', pady=(0, 10))
        
        # PDF导出按钮
        pdf_btn = tk.Button(
            export_frame,
            text="📄 PDF",
            command=self._export_security_pdf,
            font=('Microsoft YaHei UI', 8, 'bold'),
            bg='#e67e22',
            fg='white',
            activebackground='#d35400',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=8,
            borderwidth=0,
            highlightthickness=0
        )
        pdf_btn.pack(side='left', fill='x', expand=True, padx=(2, 2))
        
        # Excel导出按钮
        excel_btn = tk.Button(
            export_frame,
            text="📊 Excel",
            command=self._export_security_excel,
            font=('Microsoft YaHei UI', 8, 'bold'),
            bg='#27ae60',
            fg='white',
            activebackground='#229954',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=8,
            borderwidth=0,
            highlightthickness=0
        )
        excel_btn.pack(side='left', fill='x', expand=True, padx=(2, 2))
        
        # JSON导出按钮
        json_btn = tk.Button(
            export_frame,
            text="💾 JSON",
            command=self._export_security_json,
            font=('Microsoft YaHei UI', 8, 'bold'),
            bg='#3498db',
            fg='white',
            activebackground='#2980b9',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=8,
            borderwidth=0,
            highlightthickness=0
        )
        json_btn.pack(side='left', fill='x', expand=True, padx=(2, 2))
        
        # 悬停效果 - 增强交互
        def on_enter_pdf(e): pdf_btn.config(bg='#d35400')
        def on_leave_pdf(e): pdf_btn.config(bg='#e67e22')
        def on_click_pdf(e): pdf_btn.config(bg='#ba4a00')
        def on_enter_excel(e): excel_btn.config(bg='#229954')
        def on_leave_excel(e): excel_btn.config(bg='#27ae60')
        def on_click_excel(e): excel_btn.config(bg='#1e8449')
        def on_enter_json(e): json_btn.config(bg='#2980b9')
        def on_leave_json(e): json_btn.config(bg='#3498db')
        def on_click_json(e): json_btn.config(bg='#1f618d')
        
        pdf_btn.bind("<Enter>", on_enter_pdf)
        pdf_btn.bind("<Leave>", on_leave_pdf)
        pdf_btn.bind("<Button-1>", on_click_pdf)
        excel_btn.bind("<Enter>", on_enter_excel)
        excel_btn.bind("<Leave>", on_leave_excel)
        excel_btn.bind("<Button-1>", on_click_excel)
        json_btn.bind("<Enter>", on_enter_json)
        json_btn.bind("<Leave>", on_leave_json)
        json_btn.bind("<Button-1>", on_click_json)
        
        # 详细进度指示区 - 新增
        self.security_progress_frame = tk.Frame(content_area, bg='#ecf0f1', relief='solid', borderwidth=1)
        
        progress_container = tk.Frame(self.security_progress_frame, bg='#ecf0f1')
        progress_container.pack(fill='x', pady=8, padx=10)
        
        # 进度条 - 确定模式
        progress_bar_frame = tk.Frame(progress_container, bg='#ecf0f1')
        progress_bar_frame.pack(fill='x', pady=(0, 5))
        
        self.security_progress = ttk.Progressbar(
            progress_bar_frame,
            mode='determinate',
            length=400,
            maximum=100
        )
        self.security_progress.pack(side='left', fill='x', expand=True, padx=(0, 8))
        
        # 百分比标签
        self.security_progress_percent = tk.Label(
            progress_bar_frame,
            text="0%",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#ecf0f1',
            fg='#34495e',
            width=5
        )
        self.security_progress_percent.pack(side='left')
        
        # 进度提示标签
        self.security_progress_label = tk.Label(
            progress_container,
            text="",
            font=('Microsoft YaHei UI', 9),
            bg='#ecf0f1',
            fg='#7f8c8d',
            anchor='w'
        )
        self.security_progress_label.pack(fill='x')
        
        # 状态标签
        status_container = tk.Frame(content_area, bg='#f8f9fa', relief='solid', borderwidth=1)
        status_container.pack(fill='x', pady=(0, 10))
        
        self.security_status_label = tk.Label(
            status_container,
            text="💤 等待执行评估...",
            font=('Microsoft YaHei UI', 9),
            bg='#f8f9fa',
            fg='#95a5a6',
            pady=8
        )
        self.security_status_label.pack()
    
    def _create_actions_section(self, parent):
        """创建操作区域（垂直布局版）"""
        # 卡片外层容器 - 实现阴影效果
        shadow_frame = tk.Frame(parent, bg='#d0d8e0')
        shadow_frame.pack(fill='both', expand=True, padx=1, pady=1)
        
        # 卡片容器 - 白色背景
        card_frame = tk.Frame(shadow_frame, bg='white', relief='flat', borderwidth=0)
        card_frame.pack(fill='both', expand=True)
        
        # 卡片标题栏 - 紫色主题
        title_bar = tk.Frame(card_frame, bg='#9b59b6', height=50)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)
        
        title_icon_label = tk.Label(
            title_bar,
            text="📋",
            font=('Segoe UI Emoji', 22),
            bg='#9b59b6',
            fg='white'
        )
        title_icon_label.pack(side='left', padx=15)
        
        title_text_frame = tk.Frame(title_bar, bg='#9b59b6')
        title_text_frame.pack(side='left', fill='y', pady=8)
        
        tk.Label(
            title_text_frame,
            text="综合报告操作中心",
            font=('Microsoft YaHei UI', 12, 'bold'),
            bg='#9b59b6',
            fg='white'
        ).pack(anchor='w')
        
        tk.Label(
            title_text_frame,
            text="Comprehensive Reports",
            font=('Arial', 8),
            bg='#9b59b6',
            fg='#f3e5f5'
        ).pack(anchor='w')
        
        # 内容区域 - 优化布局
        content_area = tk.Frame(card_frame, bg='white')
        content_area.pack(fill='both', expand=True, padx=12, pady=8)
        
        # 功能说明 - 横向紧凑布局
        features_frame = tk.Frame(content_area, bg='#f3e5f5', relief='flat')
        features_frame.pack(fill='x', pady=(0, 8))
        
        features_inner = tk.Frame(features_frame, bg='#f3e5f5')
        features_inner.pack(pady=6, padx=8)
        
        features = [
            ("🔄", "完整分析"),
            ("📦", "综合PDF"),
            ("📊", "数据对比")
        ]
        
        for icon, title in features:
            item_label = tk.Label(
                features_inner,
                text=f"{icon}{title}",
                font=('Microsoft YaHei UI', 8),
                bg='#f3e5f5',
                fg='#6a1b9a',
                padx=10
            )
            item_label.pack(side='left', padx=3)
        
        # 功能介绍卡片 - 增加详细信息
        intro_frame = tk.Frame(content_area, bg='#f9f5ff', relief='flat', borderwidth=1)
        intro_frame.pack(fill='x', pady=(0, 10))
        
        tk.Label(
            intro_frame,
            text="📝 报告说明：一键执行信号分析和安全评估，\n   整合两项评估结果，生成企业级综合分析报告。",
            font=('Microsoft YaHei UI', 9),
            bg='#f9f5ff',
            fg='#6a1b9a',
            justify='left',
            pady=5,
            padx=8
        ).pack(anchor='w')
        
        # 按钮区域 - 统一布局
        button_area = tk.Frame(content_area, bg='white')
        button_area.pack(fill='both', expand=True, pady=(10, 0))
        
        # 完整分析按钮（统一高度：16px）
        complete_btn = tk.Button(
            button_area,
            text="🔄  执行完整分析",
            command=self._run_complete_analysis,
            font=('Microsoft YaHei UI', 11, 'bold'),
            bg='#9b59b6',
            fg='white',
            activebackground='#8e44ad',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=16,
            borderwidth=0,
            highlightthickness=0
        )
        complete_btn.pack(fill='x', pady=(0, 10), padx=2)
        
        # 悬停效果 - 微妙反馈
        def on_enter_complete(e): 
            complete_btn.config(bg='#8e44ad')
        def on_leave_complete(e): 
            complete_btn.config(bg='#9b59b6')
        def on_click_complete(e):
            complete_btn.config(bg='#7d3c98')
        complete_btn.bind("<Enter>", on_enter_complete)
        complete_btn.bind("<Leave>", on_leave_complete)
        complete_btn.bind("<Button-1>", on_click_complete)
        
        # 综合PDF导出按钮（统一高度：12px）
        combined_pdf_btn = tk.Button(
            button_area,
            text="📦  导出综合PDF报告",
            command=self._export_combined_pdf,
            font=('Microsoft YaHei UI', 10),
            bg='#16a085',
            fg='white',
            activebackground='#138d75',
            activeforeground='white',
            relief='flat',
            cursor='hand2',
            pady=12,
            borderwidth=0,
            highlightthickness=0
        )
        combined_pdf_btn.pack(fill='x', pady=(0, 10), padx=2)
        
        # 悬停效果 - 微妙反馈
        def on_enter_cpdf(e): 
            combined_pdf_btn.config(bg='#138d75')
        def on_leave_cpdf(e): 
            combined_pdf_btn.config(bg='#16a085')
        def on_click_cpdf(e):
            combined_pdf_btn.config(bg='#117a65')
        combined_pdf_btn.bind("<Enter>", on_enter_cpdf)
        combined_pdf_btn.bind("<Leave>", on_leave_cpdf)
        combined_pdf_btn.bind("<Button-1>", on_click_cpdf)
        
        # 详细进度指示区 - 新增
        self.complete_progress_frame = tk.Frame(content_area, bg='#ecf0f1', relief='solid', borderwidth=1)
        
        progress_container = tk.Frame(self.complete_progress_frame, bg='#ecf0f1')
        progress_container.pack(fill='x', pady=8, padx=10)
        
        # 进度条 - 确定模式
        progress_bar_frame = tk.Frame(progress_container, bg='#ecf0f1')
        progress_bar_frame.pack(fill='x', pady=(0, 5))
        
        self.complete_progress = ttk.Progressbar(
            progress_bar_frame,
            mode='determinate',
            length=400,
            maximum=100
        )
        self.complete_progress.pack(side='left', fill='x', expand=True, padx=(0, 8))
        
        # 百分比标签
        self.complete_progress_percent = tk.Label(
            progress_bar_frame,
            text="0%",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#ecf0f1',
            fg='#34495e',
            width=5
        )
        self.complete_progress_percent.pack(side='left')
        
        # 进度提示标签
        self.complete_progress_label = tk.Label(
            progress_container,
            text="",
            font=('Microsoft YaHei UI', 9),
            bg='#ecf0f1',
            fg='#7f8c8d',
            anchor='w'
        )
        self.complete_progress_label.pack(fill='x')
        
        # 功能说明提示 - 平衡布局高度
        tips_frame = tk.Frame(content_area, bg='#f3e5f5', relief='flat')
        tips_frame.pack(fill='x', pady=(10, 10))
        
        tk.Label(
            tips_frame,
            text="💡 操作流程",
            font=('Microsoft YaHei UI', 8, 'bold'),
            bg='#f3e5f5',
            fg='#6a1b9a',
            pady=3
        ).pack()
        
        tk.Label(
            tips_frame,
            text="方式一：点击'执行完整分析'一键完成\n方式二：在左侧分别执行后，点击'导出综合PDF报告'",
            font=('Microsoft YaHei UI', 9),
            bg='#f3e5f5',
            fg='#7b1fa2',
            pady=8,
            justify='left'
        ).pack()
        
        # 快速统计卡片 - 增加内容
        stats_container = tk.Frame(content_area, bg='white')
        stats_container.pack(fill='both', expand=True, pady=(0, 8))
        
        # 两列统计布局
        col1 = tk.Frame(stats_container, bg='#e8f5e9', relief='solid', borderwidth=1)
        col1.pack(side='left', fill='both', expand=True, padx=(2, 3))
        
        tk.Label(
            col1,
            text="📊",
            font=('Segoe UI Emoji', 10),
            bg='#e8f5e9'
        ).pack(pady=(3, 0))
        
        # 修改：使用实例变量以便动态更新
        self.signal_status_card = tk.Label(
            col1,
            text="信号分析\n待执行",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#e8f5e9',
            fg='#27ae60'
        )
        self.signal_status_card.pack(pady=(0, 3))
        
        col2 = tk.Frame(stats_container, bg='#ffebee', relief='solid', borderwidth=1)
        col2.pack(side='left', fill='both', expand=True, padx=(3, 2))
        
        tk.Label(
            col2,
            text="🔒",
            font=('Segoe UI Emoji', 10),
            bg='#ffebee'
        ).pack(pady=(3, 0))
        
        # 修改：使用实例变量以便动态更新
        self.security_status_card = tk.Label(
            col2,
            text="安全评估\n待执行",
            font=('Microsoft YaHei UI', 9, 'bold'),
            bg='#ffebee',
            fg='#e74c3c'
        )
        self.security_status_card.pack(pady=(0, 3))
    
    def _run_signal_analysis(self):
        """执行信号质量分析（集成详细进度显示）"""
        self.logger.info("开始执行信号质量分析")
        
        # 显示进度区域
        self.signal_progress_frame.pack(fill='x', pady=(8, 5))
        
        # 重置进度
        self.signal_progress['value'] = 0
        self.signal_progress_percent.config(text="0%")
        
        # 更新初始状态
        self._update_signal_progress(5, "🔍 初始化分析器...")
        self.signal_status_label.config(
            text="🔍 正在执行分析...",
            bg='#e3f2fd',
            fg='#2980b9',
            font=('Microsoft YaHei UI', 9, 'bold')
        )
        self.frame.update()
        
        def analysis_thread():
            try:
                # 阶段1: 获取WiFi数据
                self._update_signal_progress(10, "📡 扫描WiFi网络...")
                wifi_data = self._get_wifi_data()
                
                if not wifi_data:
                    self.logger.warning("未检测到WiFi网络")
                    self._update_signal_progress(0, "")
                    self.signal_progress_frame.pack_forget()
                    self.signal_status_label.config(
                        text="⚠️ 未检测到WiFi网络",
                        bg='#ffebee',
                        fg='#c0392b'
                    )
                    return
                
                self.logger.info(f"检测到{len(wifi_data)}个WiFi网络")
                self._update_signal_progress(25, f"✅ 发现 {len(wifi_data)} 个网络")
                
                # 阶段2: 检查缓存
                self._update_signal_progress(35, "💾 检查缓存数据...")
                cache_stats = self.cache.get_statistics()
                self.logger.debug(f"缓存统计: {cache_stats}")
                
                # 阶段3: 执行分析
                self._update_signal_progress(50, "📊 分析信号质量...")
                self.current_analysis = self.cache.get_or_compute(
                    'signal',
                    lambda: self.signal_analyzer.analyze_network_data(wifi_data)
                )
                
                # 阶段4: 计算统计指标
                self._update_signal_progress(70, "📈 计算统计指标...")
                updated_stats = self.cache.get_statistics()
                
                # 阶段5: 生成报告
                self._update_signal_progress(85, "📄 生成分析报告...")
                
                if updated_stats['signal']['hits'] > cache_stats.get('signal', {}).get('hits', 0):
                    self.logger.info("使用缓存数据，跳过重复分析")
                    cache_msg = f" (缓存有效{self.cache.signal_cache.get_remaining_time():.0f}秒)"
                else:
                    self.logger.info("执行完整信号分析")
                    cache_msg = " (已缓存5分钟)"
                
                # 完成
                self._update_signal_progress(100, "✅ 分析完成")
                quality_score = self.current_analysis.get('signal_quality', {}).get('average_score', 0)
                
                self.signal_status_label.config(
                    text=f"✅ 分析完成！质量评分: {quality_score}/100{cache_msg}",
                    bg='#e8f5e9',
                    fg='#27ae60',
                    font=('Microsoft YaHei UI', 9, 'bold')
                )
                self.logger.info(f"信号分析完成，质量评分: {quality_score}/100")
                
                # 新增：同步更新右侧综合报告中心的状态卡片
                self.signal_status_card.config(text="信号分析\n✅ 已完成", fg='#27ae60')
                
                # 启用预览按钮
                self.signal_preview_btn.config(state='normal', bg='#27ae60')
                
                # 2秒后隐藏进度条
                self.frame.after(2000, self.signal_progress_frame.pack_forget)
                
            except ValueError as e:
                self.logger.error(f"数据验证错误: {e}")
                self._update_signal_progress(0, "")
                self.signal_progress_frame.pack_forget()
                self.signal_status_label.config(
                    text=f"❌ 数据验证失败: {str(e)}",
                    bg='#fff3e0',
                    fg='#f39c12'
                )
            except Exception as e:
                self.logger.exception(f"信号分析失败: {e}")
                self._update_signal_progress(0, "")
                self.signal_progress_frame.pack_forget()
                self.signal_status_label.config(
                    text=f"❌ 分析失败: {str(e)}",
                    bg='#ffebee',
                    fg='#c0392b'
                )
        
        threading.Thread(target=analysis_thread, daemon=True).start()
    
    def _update_signal_progress(self, value, message):
        """更新信号分析进度（线程安全）"""
        def _do():
            self.signal_progress['value'] = value
            self.signal_progress_percent.config(text=f"{int(value)}%")
            self.signal_progress_label.config(text=message)
        self.frame.after(0, _do)
    
    def _run_security_assessment(self):
        """执行安全评估（集成缓存机制 + 详细进度展示）"""
        self.logger.info("开始执行PCI-DSS安全评估")
        
        def assessment_thread():
            try:
                # 显示进度区域
                self.security_progress_frame.pack(fill='x', pady=(8, 5))
                
                # 阶段1: 初始化 (5%)
                self._update_security_progress(5, "🔍 初始化安全评估器...")
                self.security_preview_btn.config(state='disabled')
                time.sleep(0.3)
                
                # 阶段2: 扫描网络 (10% → 25%)
                self._update_security_progress(10, "📡 扫描WiFi网络...")

                wifi_data = self._get_wifi_data()
                
                if not wifi_data:
                    self.logger.warning("未检测到WiFi网络")
                    self._update_security_progress(0, "")
                    self.security_progress_frame.pack_forget()
                    self.security_status_label.config(
                        text="⚠️ 未检测到WiFi网络",
                        bg='#ffebee',
                        fg='#c0392b'
                    )
                    return
                
                self._update_security_progress(25, f"✅ 发现 {len(wifi_data)} 个网络")
                self.logger.info(f"检测到{len(wifi_data)}个WiFi网络，开始安全评估")
                time.sleep(0.2)
                
                # 阶段3: 数据预处理 (35%)
                self._update_security_progress(35, "🔧 数据验证与预处理...")
                processed_data = []
                for net in wifi_data:
                    network = net.copy()
                    if 'signal_avg' not in network or network.get('signal_avg') is None:
                        signal_value = network.get('signal') or network.get('signal_percent', 0)
                        network['signal_avg'] = float(signal_value) if signal_value else 0.0
                    if 'ap_count' not in network:
                        network['ap_count'] = 1
                    processed_data.append(network)
                
                auth_ssids_text = self.auth_ssid_entry.get().strip()
                auth_ssids = [ssid.strip() for ssid in auth_ssids_text.split(',') if ssid.strip()]
                self.logger.debug(f"授权SSID列表: {auth_ssids}")
                
                # 将授权SSID转换为授权AP（使用SSID匹配）
                auth_aps = []
                for net in processed_data:
                    if net.get('ssid') in auth_ssids:
                        if net.get('bssid'):
                            auth_aps.append(net['bssid'])
                
                self.logger.debug(f"根据授权SSID匹配的AP列表: {auth_aps}")
                time.sleep(0.2)
                
                # 阶段4: 执行风险评估 (50% → 75%)
                self._update_security_progress(50, "🔒 检测加密强度...")
                cache_stats = self.cache.get_statistics()
                
                self._update_security_progress(60, "🚨 识别未授权AP...")
                time.sleep(0.3)
                
                self._update_security_progress(70, "⚠️ 配置安全审查...")
                
                self.current_assessment = self.cache.get_or_compute(
                    'security',
                    lambda: self.security_assessor.assess_network(
                        processed_data,
                        {'authorized_aps': auth_aps} if auth_aps else None
                    )
                )
                
                self._update_security_progress(75, "📊 计算风险评分...")
                time.sleep(0.2)
                
                # 阶段5: 生成报告 (85% → 100%)
                updated_stats = self.cache.get_statistics()
                if updated_stats['security']['hits'] > cache_stats.get('security', {}).get('hits', 0):
                    self.logger.info("使用缓存数据，跳过重复评估")
                    cache_msg = f" (使用缓存，剩余{self.cache.security_cache.get_remaining_time():.0f}秒)"
                else:
                    self.logger.info("执行完整安全评估")
                    cache_msg = " (已缓存10分钟)"
                
                self._update_security_progress(85, "📄 生成评估报告...")
                time.sleep(0.2)
                
                risk_score = self.current_assessment.get('overall_risk_score', 0)
                risk_level = self.current_assessment.get('risk_level', '')
                
                self._update_security_progress(100, "✅ 评估完成")
                self.logger.info(f"安全评估完成，风险评分: {risk_score}/100, 等级: {risk_level}")
                
                # 更新最终状态
                self.security_status_label.config(
                    text=f"✅ 评估完成！风险评分: {risk_score}/100 ({risk_level}){cache_msg}",
                    bg='#e8f5e9',
                    fg='#27ae60'
                )
                
                # 新增：同步更新右侧综合报告中心的状态卡片
                self.security_status_card.config(text="安全评估\n✅ 已完成", fg='#c0392b')
                
                # 启用预览按钮
                self.security_preview_btn.config(state='normal')
                
                # 2秒后隐藏进度条
                self.frame.after(2000, self.security_progress_frame.pack_forget)
                
            except ValueError as e:
                self.logger.error(f"数据验证错误: {e}")
                self._update_security_progress(0, "")
                self.security_progress_frame.pack_forget()
                self.security_status_label.config(
                    text=f"❌ 数据验证失败: {str(e)}",
                    bg='#fff3e0',
                    fg='#f39c12'
                )
            except Exception as e:
                self.logger.exception(f"安全评估失败: {e}")
                self._update_security_progress(0, "")
                self.security_progress_frame.pack_forget()
                self.security_status_label.config(
                    text=f"❌ 评估失败: {str(e)}",
                    bg='#ffebee',
                    fg='#c0392b'
                )
        
        threading.Thread(target=assessment_thread, daemon=True).start()
    
    def _update_security_progress(self, value, message):
        """更新安全评估进度（线程安全）"""
        def _do():
            self.security_progress['value'] = value
            self.security_progress_percent.config(text=f"{int(value)}%")
            self.security_progress_label.config(text=message)
        self.frame.after(0, _do)
    
    def _run_complete_analysis(self):
        """执行完整分析（带详细进度展示）"""
        def complete_thread():
            try:
                # 显示进度区域
                self.complete_progress_frame.pack(fill='x', pady=(8, 5))
                
                # 阶段1: 初始化 (5%)
                self._update_complete_progress(5, "🔍 初始化完整分析...")
                time.sleep(0.3)
                
                # 阶段2: 信号分析 (10% → 40%)
                self._update_complete_progress(10, "📡 执行信号质量分析...")
                # 更新状态卡片：进行中
                self.signal_status_card.config(text="信号分析\n进行中", fg='#f39c12')
                self._run_signal_analysis()
                
                # 等待信号分析完成（模拟）
                for i in range(3):
                    time.sleep(0.5)
                    progress = 10 + (i + 1) * 10
                    self._update_complete_progress(progress, f"📊 信号分析进行中...{i+1}/3")
                
                self._update_complete_progress(40, "✅ 信号分析完成")
                # 更新状态卡片：已完成
                self.signal_status_card.config(text="信号分析\n✅ 已完成", fg='#27ae60')
                time.sleep(0.3)
                
                # 阶段3: 安全评估 (40% → 70%)
                self._update_complete_progress(45, "🔒 执行安全风险评估...")
                # 更新状态卡片：进行中
                self.security_status_card.config(text="安全评估\n进行中", fg='#e67e22')
                self._run_security_assessment()
                
                # 等待安全评估完成（模拟）
                for i in range(3):
                    time.sleep(0.5)
                    progress = 45 + (i + 1) * 8
                    self._update_complete_progress(progress, f"🛡️ 安全评估进行中...{i+1}/3")
                
                self._update_complete_progress(70, "✅ 安全评估完成")
                # 更新状态卡片：已完成
                self.security_status_card.config(text="安全评估\n✅ 已完成", fg='#c0392b')
                time.sleep(0.3)
                
                # 阶段4: 数据整合 (70% → 85%)
                self._update_complete_progress(75, "🔄 整合分析结果...")
                time.sleep(0.5)
                self._update_complete_progress(80, "📊 生成统计图表...")
                time.sleep(0.4)
                self._update_complete_progress(85, "📝 汇总关键发现...")
                time.sleep(0.3)
                
                # 阶段5: 完成 (85% → 100%)
                self._update_complete_progress(90, "📄 准备综合报告...")
                time.sleep(0.3)
                self._update_complete_progress(95, "✅ 报告就绪...")
                time.sleep(0.2)
                self._update_complete_progress(100, "✅ 完整分析完成")
                
                self.logger.info("完整分析流程完成")
                
                # 2秒后隐藏进度条
                self.frame.after(2000, self.complete_progress_frame.pack_forget)
                
            except Exception as e:
                self.logger.exception(f"完整分析失败: {e}")
                self._update_complete_progress(0, "")
                self.complete_progress_frame.pack_forget()
                messagebox.showerror("错误", f"完整分析失败: {str(e)}")
        
        threading.Thread(target=complete_thread, daemon=True).start()
    
    def _update_complete_progress(self, value, message):
        """更新综合分析进度（线程安全）"""
        def _do():
            self.complete_progress['value'] = value
            self.complete_progress_percent.config(text=f"{int(value)}%")
            self.complete_progress_label.config(text=message)
        self.frame.after(0, _do)
    
    def _preview_signal_analysis(self):
        """预览信号分析结果"""
        if not self.current_analysis:
            messagebox.showwarning("提示", "请先执行信号质量分析")
            return
        
        # 创建预览窗口
        preview_win = tk.Toplevel(self.frame)
        preview_win.title("信号分析报告预览")
        preview_win.geometry("800x600")
        
        # 文本框显示摘要
        text_frame = ttk.Frame(preview_win)
        text_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(text_frame, wrap='word', font=('Consolas', 10))
        text_widget.pack(side='left', fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame, command=text_widget.yview)
        scrollbar.pack(side='right', fill='y')
        text_widget.config(yscrollcommand=scrollbar.set)
        
        # 生成并显示摘要
        summary = self._format_signal_analysis_summary(self.current_analysis)
        
        # 追加多点位采集数据（如果有）
        if self.collection_points:
            summary += self._get_multipoint_summary()
        
        text_widget.insert('1.0', summary)
        text_widget.config(state='disabled')
        
        # 关闭按钮
        ttk.Button(
            preview_win,
            text="关闭",
            command=preview_win.destroy
        ).pack(pady=10)
    
    def _preview_security_assessment(self):
        """预览安全评估结果"""
        if not self.current_assessment:
            messagebox.showwarning("提示", "请先执行安全评估")
            return
        
        # 创建预览窗口
        preview_win = tk.Toplevel(self.frame)
        preview_win.title("安全评估报告预览")
        preview_win.geometry("800x600")
        
        # 文本框显示报告
        text_frame = ttk.Frame(preview_win)
        text_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        text_widget = tk.Text(text_frame, wrap='word', font=('Consolas', 10))
        text_widget.pack(side='left', fill='both', expand=True)
        
        scrollbar = ttk.Scrollbar(text_frame, command=text_widget.yview)
        scrollbar.pack(side='right', fill='y')
        text_widget.config(yscrollcommand=scrollbar.set)
        
        # 显示报告
        report = self.security_assessor.generate_executive_summary()
        text_widget.insert('1.0', report)
        text_widget.config(state='disabled')
        
        # 关闭按钮
        ttk.Button(
            preview_win,
            text="关闭",
            command=preview_win.destroy
        ).pack(pady=10)
    
    def _export_signal_pdf(self):
        """导出信号分析PDF"""
        if not self.current_analysis:
            messagebox.showwarning("提示", "请先执行信号质量分析")
            return
        
        # 选择保存路径
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"WiFi信号分析报告_{timestamp}.pdf"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF文件", "*.pdf")],
            initialfile=default_name
        )
        
        if filepath:
            try:
                success = self.report_generator.generate_report(
                    self.current_analysis,
                    filepath,
                    self.signal_template,
                    company_name="企业名称",
                    report_type="signal"
                )
                
                if success:
                    messagebox.showinfo("成功", f"PDF报告已保存到:\n{filepath}")
                else:
                    messagebox.showerror("错误", "PDF报告生成失败")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def _export_signal_json(self):
        """导出信号分析JSON"""
        if not self.current_analysis:
            messagebox.showwarning("提示", "请先执行信号质量分析")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"WiFi信号分析数据_{timestamp}.json"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON文件", "*.json")],
            initialfile=default_name
        )
        
        if filepath:
            try:
                success = self.signal_analyzer.export_to_json(filepath)
                if success:
                    messagebox.showinfo("成功", f"JSON数据已保存到:\n{filepath}")
                else:
                    messagebox.showerror("错误", "JSON导出失败")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def _export_signal_excel(self):
        """导出信号分析数据到Excel"""
        if not self.current_analysis:
            messagebox.showwarning("提示", "请先执行信号质量分析")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"WiFi信号分析数据_{timestamp}.xlsx"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=default_name
        )
        
        if filepath:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                wb = Workbook()
                
                # 1. 概览工作表
                ws_summary = wb.active
                ws_summary.title = "分析概览"
                
                # 标题
                ws_summary['A1'] = "WiFi信号质量分析报告"
                ws_summary['A1'].font = Font(size=16, bold=True, color="1F4E78")
                ws_summary.merge_cells('A1:D1')
                
                # 基本信息
                row = 3
                quality_data = self.current_analysis.get('signal_quality', {})
                ws_summary[f'A{row}'] = "分析时间"
                ws_summary[f'B{row}'] = self.current_analysis.get('scan_time', '')
                row += 1
                ws_summary[f'A{row}'] = "网络总数"
                ws_summary[f'B{row}'] = quality_data.get('total_networks', 0)
                row += 1
                ws_summary[f'A{row}'] = "平均信号强度"
                ws_summary[f'B{row}'] = f"{quality_data.get('average_signal', 0):.1f}%"
                row += 1
                ws_summary[f'A{row}'] = "平均dBm值"
                ws_summary[f'B{row}'] = f"{quality_data.get('average_dbm', -100):.1f} dBm"
                row += 1
                ws_summary[f'A{row}'] = "质量评分"
                ws_summary[f'B{row}'] = f"{quality_data.get('average_score', 0)}/100"
                row += 1
                ws_summary[f'A{row}'] = "质量评级"
                ws_summary[f'B{row}'] = quality_data.get('quality_rating', '未知')
                
                # 信号分布统计
                row += 2
                ws_summary[f'A{row}'] = "信号分布统计"
                ws_summary[f'A{row}'].font = Font(size=12, bold=True)
                row += 1
                ws_summary[f'A{row}'] = "优秀(≥80%)"
                ws_summary[f'B{row}'] = quality_data.get('excellent_count', 0)
                row += 1
                ws_summary[f'A{row}'] = "良好(60-79%)"
                ws_summary[f'B{row}'] = quality_data.get('good_count', 0)
                row += 1
                ws_summary[f'A{row}'] = "一般(40-59%)"
                ws_summary[f'B{row}'] = quality_data.get('fair_count', 0)
                row += 1
                ws_summary[f'A{row}'] = "较差(<40%)"
                ws_summary[f'B{row}'] = quality_data.get('poor_count', 0)
                
                # 2. 网络详情工作表
                ws_networks = wb.create_sheet("网络详情")
                headers = ["SSID", "信号强度(%)", "dBm值", "加密方式", "信道", "频段", "AP数量"]
                for col, header in enumerate(headers, 1):
                    cell = ws_networks.cell(1, col, header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                networks = self.current_analysis.get('networks', [])
                for row, net in enumerate(networks, 2):
                    ws_networks.cell(row, 1, net.get('ssid', ''))
                    signal_percent = net.get('signal_percent', net.get('signal', 0))
                    ws_networks.cell(row, 2, f"{signal_percent:.1f}")
                    # 计算dBm
                    dbm_value = (signal_percent * 0.7) - 100 if 0 <= signal_percent <= 100 else signal_percent
                    ws_networks.cell(row, 3, f"{dbm_value:.1f}")
                    ws_networks.cell(row, 4, net.get('encryption', ''))
                    ws_networks.cell(row, 5, net.get('channel', ''))
                    ws_networks.cell(row, 6, net.get('band', ''))
                    ws_networks.cell(row, 7, net.get('ap_count', 1))
                
                # 3. 多点位采集数据（如果有）
                if self.collection_points:
                    ws_multipoint = wb.create_sheet("多点位数据")
                    headers = ["点位名称", "采集时间", "网络数", "平均信号(%)", "最强信号(%)", "最弱信号(%)", "强信号网络数"]
                    for col, header in enumerate(headers, 1):
                        cell = ws_multipoint.cell(1, col, header)
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center")
                    
                    for row, point in enumerate(self.collection_points, 2):
                        ws_multipoint.cell(row, 1, point['location'])
                        ws_multipoint.cell(row, 2, point['timestamp'])
                        ws_multipoint.cell(row, 3, point['network_count'])
                        
                        if point['networks']:
                            signals = [net.get('signal_percent', 0) for net in point['networks']]
                            ws_multipoint.cell(row, 4, f"{sum(signals)/len(signals):.1f}")
                            ws_multipoint.cell(row, 5, f"{max(signals):.1f}")
                            ws_multipoint.cell(row, 6, f"{min(signals):.1f}")
                            ws_multipoint.cell(row, 7, sum(1 for s in signals if s > 70))
                
                # 调整列宽
                for ws in wb.worksheets:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filepath)
                self.logger.info(f"Excel数据已导出到: {filepath}")
                messagebox.showinfo("成功", f"Excel文件已保存到:\n{filepath}")
                
            except ImportError:
                messagebox.showerror("错误", "Excel导出功能需要安装openpyxl库\n请运行: pip install openpyxl")
            except Exception as e:
                self.logger.error(f"Excel导出失败: {e}")
                messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def _export_security_excel(self):
        """导出安全评估数据到Excel"""
        if not self.current_assessment:
            messagebox.showwarning("提示", "请先执行安全评估")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"PCI-DSS安全评估数据_{timestamp}.xlsx"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=default_name
        )
        
        if filepath:
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
                
                wb = Workbook()
                
                # 1. 评估概览工作表
                ws_summary = wb.active
                ws_summary.title = "评估概览"
                
                # 标题
                ws_summary['A1'] = "PCI-DSS无线网络安全评估报告"
                ws_summary['A1'].font = Font(size=16, bold=True, color="C0392B")
                ws_summary.merge_cells('A1:D1')
                
                # 基本信息
                row = 3
                ws_summary[f'A{row}'] = "评估时间"
                ws_summary[f'B{row}'] = self.current_assessment.get('scan_time', '')
                row += 1
                ws_summary[f'A{row}'] = "扫描网络数"
                ws_summary[f'B{row}'] = len(self.current_assessment.get('networks', []))
                row += 1
                ws_summary[f'A{row}'] = "风险等级"
                ws_summary[f'B{row}'] = self.current_assessment.get('risk_level', '未知')
                row += 1
                ws_summary[f'A{row}'] = "合规状态"
                ws_summary[f'B{row}'] = self.current_assessment.get('compliance_status', '未知')
                row += 1
                ws_summary[f'A{row}'] = "安全评分"
                ws_summary[f'B{row}'] = f"{self.current_assessment.get('security_score', 0)}/100"
                
                # 风险统计
                row += 2
                ws_summary[f'A{row}'] = "风险统计"
                ws_summary[f'A{row}'].font = Font(size=12, bold=True)
                row += 1
                
                risk_summary = self.current_assessment.get('risk_summary', {})
                ws_summary[f'A{row}'] = "高风险网络"
                ws_summary[f'B{row}'] = risk_summary.get('high_risk', 0)
                row += 1
                ws_summary[f'A{row}'] = "中风险网络"
                ws_summary[f'B{row}'] = risk_summary.get('medium_risk', 0)
                row += 1
                ws_summary[f'A{row}'] = "低风险网络"
                ws_summary[f'B{row}'] = risk_summary.get('low_risk', 0)
                row += 1
                ws_summary[f'A{row}'] = "未授权AP"
                ws_summary[f'B{row}'] = risk_summary.get('unauthorized_aps', 0)
                row += 1
                ws_summary[f'A{row}'] = "弱加密网络"
                ws_summary[f'B{row}'] = risk_summary.get('weak_encryption', 0)
                
                # 授权SSID列表
                if self.saved_auth_ssids:
                    row += 2
                    ws_summary[f'A{row}'] = "授权WiFi SSID"
                    ws_summary[f'A{row}'].font = Font(size=12, bold=True)
                    row += 1
                    for ssid in self.saved_auth_ssids:
                        ws_summary[f'A{row}'] = ssid
                        ws_summary[f'B{row}'] = "✓ 授权"
                        row += 1
                
                # 2. 网络详情工作表
                ws_networks = wb.create_sheet("网络详情")
                headers = ["SSID", "BSSID", "加密方式", "信号强度(%)", "信道", "频段", "风险等级", "风险原因", "合规状态"]
                for col, header in enumerate(headers, 1):
                    cell = ws_networks.cell(1, col, header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                networks = self.current_assessment.get('networks', [])
                for row, net in enumerate(networks, 2):
                    ws_networks.cell(row, 1, net.get('ssid', ''))
                    ws_networks.cell(row, 2, net.get('bssid', ''))
                    ws_networks.cell(row, 3, net.get('encryption', ''))
                    ws_networks.cell(row, 4, f"{net.get('signal_percent', 0):.1f}")
                    ws_networks.cell(row, 5, net.get('channel', ''))
                    ws_networks.cell(row, 6, net.get('band', ''))
                    
                    # 风险评估信息
                    risk_level = net.get('risk_level', '未知')
                    ws_networks.cell(row, 7, risk_level)
                    
                    # 根据风险等级设置背景色
                    risk_cell = ws_networks.cell(row, 7)
                    if risk_level == '高风险':
                        risk_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                    elif risk_level == '中风险':
                        risk_cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
                    elif risk_level == '低风险':
                        risk_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                    
                    ws_networks.cell(row, 8, ', '.join(net.get('risk_reasons', [])))
                    ws_networks.cell(row, 9, '✓ 合规' if net.get('is_compliant', False) else '✗ 不合规')
                
                # 3. 风险详情工作表
                ws_risks = wb.create_sheet("风险详情")
                headers = ["风险类型", "影响网络", "严重程度", "建议措施"]
                for col, header in enumerate(headers, 1):
                    cell = ws_risks.cell(1, col, header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # 提取所有风险
                row = 2
                for net in networks:
                    if net.get('risk_reasons'):
                        for risk in net['risk_reasons']:
                            ws_risks.cell(row, 1, risk)
                            ws_risks.cell(row, 2, net.get('ssid', ''))
                            ws_risks.cell(row, 3, net.get('risk_level', ''))
                            
                            # 根据风险类型提供建议
                            if '未授权' in risk:
                                suggestion = "阻止未授权AP接入，加强物理安全控制"
                            elif '加密' in risk or 'WEP' in risk or 'WPA' in risk:
                                suggestion = "升级到WPA3加密协议，禁用弱加密方式"
                            elif 'WPS' in risk:
                                suggestion = "禁用WPS功能，使用强密码策略"
                            elif '信道' in risk:
                                suggestion = "优化信道配置，减少干扰和冲突"
                            else:
                                suggestion = "评估安全配置，遵循PCI-DSS标准"
                            
                            ws_risks.cell(row, 4, suggestion)
                            row += 1
                
                # 4. 合规性检查工作表
                ws_compliance = wb.create_sheet("合规性检查")
                headers = ["检查项目", "要求", "当前状态", "合规性"]
                for col, header in enumerate(headers, 1):
                    cell = ws_compliance.cell(1, col, header)
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # 合规性检查项
                compliance_items = [
                    ("加密强度", "所有无线网络必须使用WPA2/WPA3加密", 
                     f"{sum(1 for n in networks if 'WPA2' in n.get('encryption', '') or 'WPA3' in n.get('encryption', ''))}/{len(networks)} 网络符合",
                     "✓" if all('WPA2' in n.get('encryption', '') or 'WPA3' in n.get('encryption', '') or not n.get('encryption') for n in networks) else "✗"),
                    ("授权管理", "仅允许授权的WiFi SSID接入",
                     f"已配置 {len(self.saved_auth_ssids)} 个授权SSID" if self.saved_auth_ssids else "未配置授权SSID",
                     "✓" if self.saved_auth_ssids else "⚠"),
                    ("WPS安全", "禁用WPS功能防止暴力破解",
                     "需手动验证AP配置",
                     "⚠"),
                    ("信号隔离", "确保无线网络与支付系统网络隔离",
                     "需手动验证网络架构",
                     "⚠"),
                    ("定期审计", "每季度进行安全评估和渗透测试",
                     f"本次评估: {self.current_assessment.get('scan_time', '')}",
                     "⚠")
                ]
                
                for row, (item, requirement, status, compliance) in enumerate(compliance_items, 2):
                    ws_compliance.cell(row, 1, item)
                    ws_compliance.cell(row, 2, requirement)
                    ws_compliance.cell(row, 3, status)
                    ws_compliance.cell(row, 4, compliance)
                    
                    # 设置合规性单元格颜色
                    comp_cell = ws_compliance.cell(row, 4)
                    if compliance == "✓":
                        comp_cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
                        comp_cell.font = Font(color="27AE60", bold=True)
                    elif compliance == "✗":
                        comp_cell.fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
                        comp_cell.font = Font(color="C0392B", bold=True)
                    else:
                        comp_cell.fill = PatternFill(start_color="FFF4CC", end_color="FFF4CC", fill_type="solid")
                        comp_cell.font = Font(color="E67E22", bold=True)
                
                # 调整所有工作表的列宽
                for ws in wb.worksheets:
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                
                wb.save(filepath)
                self.logger.info(f"安全评估Excel数据已导出到: {filepath}")
                messagebox.showinfo("成功", f"Excel文件已保存到:\n{filepath}")
                
            except ImportError:
                messagebox.showerror("错误", "Excel导出功能需要安装openpyxl库\n请运行: pip install openpyxl")
            except Exception as e:
                self.logger.error(f"安全评估Excel导出失败: {e}")
                messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def _export_security_pdf(self):
        """导出安全评估PDF"""
        if not self.current_assessment:
            messagebox.showwarning("提示", "请先执行安全评估")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"PCI-DSS安全评估报告_{timestamp}.pdf"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF文件", "*.pdf")],
            initialfile=default_name
        )
        
        if filepath:
            try:
                success = self.report_generator.generate_report(
                    self.current_assessment,
                    filepath,
                    self.pci_template,
                    company_name="企业名称",
                    report_type="pci_dss"
                )
                
                if success:
                    messagebox.showinfo("成功", f"PDF报告已保存到:\n{filepath}")
                else:
                    messagebox.showerror("错误", "PDF报告生成失败")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def _export_security_json(self):
        """导出安全评估JSON"""
        if not self.current_assessment:
            messagebox.showwarning("提示", "请先执行安全评估")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"PCI-DSS安全评估数据_{timestamp}.json"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".json",
            filetypes=[("JSON文件", "*.json")],
            initialfile=default_name
        )
        
        if filepath:
            try:
                success = self.security_assessor.export_to_json(filepath)
                if success:
                    messagebox.showinfo("成功", f"JSON数据已保存到:\n{filepath}")
                else:
                    messagebox.showerror("错误", "JSON导出失败")
            except Exception as e:
                messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def _export_combined_pdf(self):
        """导出综合PDF报告"""
        self.logger.info("用户请求导出综合PDF报告")
        
        if not self.current_analysis and not self.current_assessment:
            self.logger.warning("导出失败：未执行任何分析")
            messagebox.showwarning("提示", "请先执行至少一项分析")
            return
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_name = f"WiFi综合分析报告_{timestamp}.pdf"
        
        filepath = filedialog.asksaveasfilename(
            defaultextension=".pdf",
            filetypes=[("PDF文件", "*.pdf")],
            initialfile=default_name
        )
        
        if filepath:
            try:
                self.logger.info(f"开始生成PDF报告: {filepath}")
                
                # 根据可用数据生成报告
                if self.current_analysis and self.current_assessment:
                    # 生成两份报告并合并（简化版：分别生成）
                    base_path = filepath.replace('.pdf', '')
                    signal_path = f"{base_path}_信号分析.pdf"
                    security_path = f"{base_path}_安全评估.pdf"
                    
                    self.logger.info("生成信号分析报告")
                    self.report_generator.generate_report(
                        self.current_analysis, signal_path, self.signal_template,
                        company_name="企业名称", report_type="signal"
                    )
                    
                    self.logger.info("生成PCI-DSS安全评估报告")
                    self.report_generator.generate_report(
                        self.current_assessment, security_path, self.pci_template,
                        company_name="企业名称", report_type="pci_dss"
                    )
                    
                    self.logger.info(f"综合报告生成成功: {signal_path}, {security_path}")
                    messagebox.showinfo(
                        "成功",
                        f"综合报告已分别保存到:\n{signal_path}\n{security_path}"
                    )
                    
                elif self.current_analysis:
                    self.logger.info("仅生成信号分析报告")
                    success = self.report_generator.generate_report(
                        self.current_analysis, filepath, self.signal_template,
                        company_name="企业名称", report_type="signal"
                    )
                    if success:
                        self.logger.info(f"信号分析报告生成成功: {filepath}")
                        messagebox.showinfo("成功", f"信号分析报告已保存到:\n{filepath}")
                        
                elif self.current_assessment:
                    self.logger.info("仅生成安全评估报告")
                    success = self.report_generator.generate_report(
                        self.current_assessment, filepath, self.pci_template,
                        company_name="企业名称", report_type="pci_dss"
                    )
                    if success:
                        self.logger.info(f"安全评估报告生成成功: {filepath}")
                        messagebox.showinfo("成功", f"安全评估报告已保存到:\n{filepath}")
                        
            except PermissionError:
                self.logger.error(f"权限错误：无法写入文件 {filepath}")
                messagebox.showerror("错误", f"无法写入文件，请检查文件是否被占用或权限不足")
            except IOError as e:
                self.logger.error(f"IO错误: {e}")
                messagebox.showerror("错误", f"文件操作失败: {str(e)}")
            except Exception as e:
                self.logger.exception(f"PDF导出失败: {e}")
                messagebox.showerror("错误", f"导出失败: {str(e)}")
    
    def _get_wifi_data(self) -> list:
        """获取WiFi扫描数据"""
        try:
            self.logger.info("开始扫描WiFi网络")
            # 扫描WiFi网络（使用正确的方法名）
            networks = self.wifi_analyzer.scan_wifi_networks(force_refresh=True)
            
            if not networks:
                self.logger.warning("WiFi扫描未返回任何网络")
                return []
            
            self.logger.info(f"成功扫描到{len(networks)}个WiFi网络")
            
            # 转换为标准格式
            wifi_data = []
            for network in networks:
                # 修复：使用 signal_percent（整数百分比）而不是 signal
                # scan_wifi_networks 返回的数据包含:
                # - signal_strength: 字符串形式（如"75%"）
                # - signal_percent: 整数形式（如75）
                signal_value = network.get('signal_percent', network.get('signal', 0))
                
                wifi_data.append({
                    'ssid': network.get('ssid', ''),
                    'bssid': network.get('bssid', ''),
                    'signal': signal_value,  # 使用整数百分比值（0-100）
                    'channel': network.get('channel', 0),
                    'authentication': network.get('authentication', ''),
                    'encryption': network.get('encryption', ''),
                    'band': network.get('band', '')  # 添加频段信息
                })
            
            return wifi_data
            
        except subprocess.CalledProcessError as e:
            self.logger.error(f"netsh命令执行失败: {e.cmd}, 返回码: {e.returncode}")
            return []
        except FileNotFoundError:
            self.logger.error("netsh命令不可用，请检查系统环境")
            return []
        except Exception as e:
            self.logger.exception(f"获取WiFi数据失败: {e}")
            return []
    
    def _format_signal_analysis_summary(self, analysis_data: dict) -> str:
        """格式化信号分析摘要"""
        summary = []
        summary.append("="*60)
        summary.append("        WiFi信号质量分析报告")
        summary.append("="*60)
        summary.append("")
        
        # 基本信息
        summary.append(f"分析时间: {analysis_data.get('scan_time', 'N/A')}")
        summary.append(f"扫描网络数量: {analysis_data.get('total_networks', 0)}")
        summary.append(f"扫描次数: {analysis_data.get('scan_count', 1)}")
        summary.append("")
        
        # 信号质量分析
        signal_quality = analysis_data.get('signal_quality', {})
        if signal_quality:
            summary.append("-"*60)
            summary.append("1. 信号质量分析")
            summary.append("-"*60)
            # 修复：添加正确的百分比单位，并显示dBm参考值
            avg_signal_pct = signal_quality.get('average_signal', 0)
            avg_dbm = signal_quality.get('average_dbm', -100)
            summary.append(f"平均信号强度: {avg_signal_pct:.1f}% ({avg_dbm:.1f} dBm)")
            summary.append(f"质量评级: {signal_quality.get('average_quality', 'N/A')}")
            summary.append(f"最强信号: {signal_quality.get('max_signal', 0):.0f}%")
            summary.append(f"最弱信号: {signal_quality.get('min_signal', 0):.0f}%")
            summary.append("")
            summary.append("信号分布:")
            summary.append(f"  优秀 (≥80%): {signal_quality.get('excellent_count', 0)} 个")
            summary.append(f"  良好 (60-79%): {signal_quality.get('good_count', 0)} 个")
            summary.append(f"  一般 (40-59%): {signal_quality.get('fair_count', 0)} 个")
            summary.append(f"  较差 (<40): {signal_quality.get('poor_count', 0)} 个")
            summary.append("")
        
        # 覆盖分析
        coverage = analysis_data.get('coverage', {})
        if coverage:
            summary.append("-"*60)
            summary.append("2. 覆盖范围分析")
            summary.append("-"*60)
            summary.append(f"总接入点数: {coverage.get('total_aps', 0)}")
            summary.append(f"唯一SSID数: {coverage.get('unique_ssids', 0)}")
            summary.append(f"覆盖评级: {coverage.get('coverage_rating', 'N/A')}")
            summary.append("")
            summary.append("频段分布:")
            summary.append(f"  2.4GHz: {coverage.get('2.4ghz_count', 0)} 个 ({coverage.get('frequency_distribution', {}).get('2.4GHz', 0):.1f}%)")
            summary.append(f"  5GHz: {coverage.get('5ghz_count', 0)} 个 ({coverage.get('frequency_distribution', {}).get('5GHz', 0):.1f}%)")
            summary.append("")
        
        # 干扰分析
        interference = analysis_data.get('interference', {})
        if interference:
            summary.append("-"*60)
            summary.append("3. 干扰分析")
            summary.append("-"*60)
            summary.append(f"使用信道数: {interference.get('total_channels_used', 0)}")
            summary.append(f"干扰级别: {interference.get('interference_level', 'N/A')}")
            summary.append("")
            crowded = interference.get('most_crowded_channels', [])
            if crowded:
                summary.append("最拥挤信道:")
                for ch in crowded[:5]:
                    summary.append(f"  信道 {ch.get('channel', 0)}: {ch.get('ap_count', 0)} 个AP")
            summary.append("")
        
        # 信道使用
        channel_usage = analysis_data.get('channel_usage', {})
        if channel_usage:
            summary.append("-"*60)
            summary.append("4. 信道使用情况")
            summary.append("-"*60)
            
            ch_24 = channel_usage.get('2.4GHz', {})
            if ch_24:
                summary.append("2.4GHz频段:")
                for ch, count in sorted(ch_24.items(), key=lambda x: x[1] if isinstance(x[1], int) else len(x[1]), reverse=True)[:5]:
                    ap_count = count if isinstance(count, int) else len(count)
                    summary.append(f"  信道 {ch}: {ap_count} 个AP")
                if ch_24.get('recommended_channel'):
                    summary.append(f"  推荐信道: {ch_24['recommended_channel']}")
                summary.append("")
            
            ch_5 = channel_usage.get('5GHz', {})
            if ch_5:
                summary.append("5GHz频段:")
                for ch, count in sorted(ch_5.items(), key=lambda x: x[1] if isinstance(x[1], int) else len(x[1]), reverse=True)[:5]:
                    ap_count = count if isinstance(count, int) else len(count)
                    summary.append(f"  信道 {ch}: {ap_count} 个AP")
                if ch_5.get('recommended_channel'):
                    summary.append(f"  推荐信道: {ch_5['recommended_channel']}")
                summary.append("")
        
        # 优化建议
        recommendations = analysis_data.get('recommendations', [])
        if recommendations:
            summary.append("-"*60)
            summary.append("5. 优化建议")
            summary.append("-"*60)
            for i, rec in enumerate(recommendations, 1):
                summary.append(f"{i}. {rec}")
            summary.append("")
        
        summary.append("="*60)
        
        return "\n".join(summary)
    
    def _add_collection_point(self):
        """添加一个信号采集点"""
        location_name = self.location_entry.get().strip()
        
        # 验证位置名称
        if not location_name or location_name == "办公区A-1层":
            messagebox.showwarning("提示", "请输入有效的位置名称")
            return
        
        # 检查是否重复
        if any(p['location'] == location_name for p in self.collection_points):
            messagebox.showwarning("提示", f"位置 '{location_name}' 已存在，请使用不同名称")
            return
        
        # 显示采集进度
        self.logger.info(f"开始采集点位: {location_name}")
        
        # 扫描WiFi网络
        wifi_data = self._get_wifi_data()
        
        if not wifi_data:
            messagebox.showerror("错误", "未检测到WiFi网络，请检查WiFi是否开启")
            return
        
        # 保存采集数据
        collection_data = {
            'location': location_name,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'networks': wifi_data,
            'network_count': len(wifi_data)
        }
        
        self.collection_points.append(collection_data)
        
        # 更新显示
        self.points_label.config(text=f"已采集: {len(self.collection_points)} 点")
        
        # 清空输入框
        self.location_entry.delete(0, tk.END)
        self.location_entry.insert(0, f"办公区-点{len(self.collection_points) + 1}")
        
        # 日志记录
        self.logger.info(f"采集点 '{location_name}' 添加成功，检测到 {len(wifi_data)} 个网络")
        
        messagebox.showinfo(
            "采集成功",
            f"位置: {location_name}\n"
            f"时间: {collection_data['timestamp']}\n"
            f"网络数: {len(wifi_data)} 个\n\n"
            f"总采集点: {len(self.collection_points)} 个"
        )
    
    def _get_multipoint_summary(self):
        """获取多点位采集数据摘要"""
        if not self.collection_points:
            return "\n" + "="*60 + "\n未进行多点位采集\n" + "="*60 + "\n"
        
        summary = ["\n" + "="*60]
        summary.append("多点位信号采集数据")
        summary.append("="*60)
        
        for i, point in enumerate(self.collection_points, 1):
            summary.append(f"\n点位 {i}: {point['location']}")
            summary.append(f"采集时间: {point['timestamp']}")
            summary.append(f"检测网络: {point['network_count']} 个")
            
            # 统计信号强度分布
            if point['networks']:
                signals = [net.get('signal_percent', 0) for net in point['networks']]
                avg_signal = sum(signals) / len(signals) if signals else 0
                max_signal = max(signals) if signals else 0
                min_signal = min(signals) if signals else 0
                
                summary.append(f"平均信号: {avg_signal:.1f}%")
                summary.append(f"最强信号: {max_signal:.1f}%")
                summary.append(f"最弱信号: {min_signal:.1f}%")
                
                # 统计强信号网络（>70%）
                strong_count = sum(1 for s in signals if s > 70)
                summary.append(f"强信号网络(>70%): {strong_count} 个")
        
        summary.append("\n" + "="*60)
        summary.append(f"总采集点位: {len(self.collection_points)} 个")
        summary.append("="*60 + "\n")
        
        return "\n".join(summary)
    
    def _save_auth_ssids(self):
        """保存授权SSID到本地配置文件"""
        try:
            auth_ssids_text = self.auth_ssid_entry.get().strip()
            if not auth_ssids_text:
                messagebox.showwarning("提示", "授权SSID列表为空")
                return
            
            auth_ssids = [s.strip() for s in auth_ssids_text.split(',') if s.strip()]
            
            with open(self.auth_ssid_file, 'w', encoding='utf-8') as f:
                import json
                json.dump({'ssids': auth_ssids, 'updated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S')}, f, ensure_ascii=False, indent=2)
            
            self.saved_auth_ssids = auth_ssids
            self.logger.info(f"授权SSID已保存: {auth_ssids}")
            messagebox.showinfo("成功", f"已保存 {len(auth_ssids)} 个授权SSID")
            
        except Exception as e:
            self.logger.error(f"保存授权SSID失败: {e}")
            messagebox.showerror("错误", f"保存失败: {str(e)}")
    
    def _load_auth_ssids(self):
        """从本地配置文件加载授权SSID"""
        try:
            if os.path.exists(self.auth_ssid_file):
                with open(self.auth_ssid_file, 'r', encoding='utf-8') as f:
                    import json
                    data = json.load(f)
                    ssids = data.get('ssids', [])
                    self.logger.info(f"已加载授权SSID: {ssids}")
                    return ssids
        except Exception as e:
            self.logger.error(f"加载授权SSID失败: {e}")
        return []
    
    def _save_multipoint_data(self):
        """保存多点位采集数据到文件"""
        try:
            if not self.collection_points:
                messagebox.showwarning("提示", "没有采集数据可保存")
                return
            
            # 选择保存路径
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            default_name = f"多点位采集数据_{timestamp}.json"
            
            filepath = filedialog.asksaveasfilename(
                defaultextension=".json",
                filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")],
                initialfile=default_name,
                initialdir=os.path.dirname(self.multipoint_file)
            )
            
            if filepath:
                with open(filepath, 'w', encoding='utf-8') as f:
                    import json
                    json.dump({
                        'saved_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        'total_points': len(self.collection_points),
                        'collections': self.collection_points
                    }, f, ensure_ascii=False, indent=2)
                
                self.logger.info(f"多点位数据已保存到: {filepath}")
                messagebox.showinfo("成功", f"已保存 {len(self.collection_points)} 个采集点数据到:\n{os.path.basename(filepath)}")
        
        except Exception as e:
            self.logger.error(f"保存多点位数据失败: {e}")
            messagebox.showerror("错误", f"保存失败: {str(e)}")
    
    def _load_multipoint_data(self):
        """从默认配置文件加载多点位数据（自动加载）"""
        try:
            if os.path.exists(self.multipoint_file):
                with open(self.multipoint_file, 'r', encoding='utf-8') as f:
                    import json
                    data = json.load(f)
                    collections = data.get('collections', [])
                    if collections:
                        self.collection_points = collections
                        self.logger.info(f"已自动加载 {len(collections)} 个采集点")
        except Exception as e:
            self.logger.error(f"自动加载多点位数据失败: {e}")
    
    def _load_multipoint_data_from_file(self):
        """从用户选择的文件加载多点位数据"""
        try:
            filepath = filedialog.askopenfilename(
                title="选择多点位采集数据文件",
                filetypes=[("JSON文件", "*.json"), ("所有文件", "*.*")],
                initialdir=os.path.dirname(self.multipoint_file)
            )
            
            if filepath:
                with open(filepath, 'r', encoding='utf-8') as f:
                    import json
                    data = json.load(f)
                    collections = data.get('collections', [])
                    
                    if collections:
                        # 询问是追加还是替换
                        if self.collection_points:
                            choice = messagebox.askyesnocancel(
                                "加载方式",
                                f"当前已有 {len(self.collection_points)} 个采集点\n\n"
                                f"是(Yes): 追加到现有数据\n"
                                f"否(No): 替换现有数据\n"
                                f"取消: 不加载"
                            )
                            if choice is None:  # 取消
                                return
                            elif choice:  # 追加
                                self.collection_points.extend(collections)
                            else:  # 替换
                                self.collection_points = collections
                        else:
                            self.collection_points = collections
                        
                        # 更新显示
                        self.points_label.config(text=f"已采集: {len(self.collection_points)} 点")
                        
                        self.logger.info(f"已加载 {len(collections)} 个采集点从: {filepath}")
                        messagebox.showinfo("成功", f"已加载 {len(collections)} 个采集点\n总计: {len(self.collection_points)} 点")
                    else:
                        messagebox.showwarning("提示", "文件中没有采集数据")
        
        except Exception as e:
            self.logger.error(f"加载多点位数据失败: {e}")
            messagebox.showerror("错误", f"加载失败: {str(e)}")
    
    def _clear_multipoint_data(self):
        """清空多点位采集数据"""
        if not self.collection_points:
            messagebox.showinfo("提示", "当前没有采集数据")
            return
        
        if messagebox.askyesno("确认", f"确定要清空所有 {len(self.collection_points)} 个采集点的数据吗？\n此操作不可恢复！"):
            self.collection_points = []
            self.points_label.config(text="已采集: 0 点")
            self.logger.info("多点位采集数据已清空")
            messagebox.showinfo("成功", "已清空所有采集数据")
    
    def _generate_multipoint_chart(self):
        """生成多点位对比图表"""
        if not self.collection_points:
            messagebox.showwarning("提示", "请先采集至少一个点位的数据")
            return
        
        if len(self.collection_points) < 2:
            messagebox.showwarning("提示", "至少需要2个点位数据才能生成对比图表")
            return
        
        try:
            import matplotlib.pyplot as plt
            import matplotlib
            from matplotlib.font_manager import FontProperties
            
            # 设置中文字体
            matplotlib.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
            matplotlib.rcParams['axes.unicode_minus'] = False
            
            # 准备数据
            locations = [p['location'] for p in self.collection_points]
            avg_signals = []
            max_signals = []
            min_signals = []
            strong_counts = []
            network_counts = []
            
            for point in self.collection_points:
                if point['networks']:
                    signals = [net.get('signal_percent', 0) for net in point['networks']]
                    avg_signals.append(sum(signals) / len(signals))
                    max_signals.append(max(signals))
                    min_signals.append(min(signals))
                    strong_counts.append(sum(1 for s in signals if s > 70))
                    network_counts.append(len(signals))
                else:
                    avg_signals.append(0)
                    max_signals.append(0)
                    min_signals.append(0)
                    strong_counts.append(0)
                    network_counts.append(0)
            
            # 创建图表 (2行2列)
            fig, ((ax1, ax2), (ax3, ax4)) = plt.subplots(2, 2, figsize=(14, 10))
            fig.suptitle('多点位WiFi信号对比分析', fontsize=16, fontweight='bold')
            
            x_pos = range(len(locations))
            
            # 图1: 信号强度对比（柱状图）
            bars1 = ax1.bar(x_pos, avg_signals, color='#3498db', alpha=0.7, label='平均信号')
            ax1.plot(x_pos, max_signals, 'ro-', label='最强信号', linewidth=2)
            ax1.plot(x_pos, min_signals, 'gs-', label='最弱信号', linewidth=2)
            ax1.set_xlabel('采集点位', fontsize=10)
            ax1.set_ylabel('信号强度 (%)', fontsize=10)
            ax1.set_title('信号强度对比', fontsize=12, fontweight='bold')
            ax1.set_xticks(x_pos)
            ax1.set_xticklabels(locations, rotation=15, ha='right', fontsize=9)
            ax1.legend(fontsize=9)
            ax1.grid(axis='y', alpha=0.3)
            ax1.axhline(y=70, color='orange', linestyle='--', alpha=0.5, label='优良线(70%)')
            
            # 在柱状图上显示数值
            for i, (bar, val) in enumerate(zip(bars1, avg_signals)):
                ax1.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 2, 
                        f'{val:.1f}%', ha='center', va='bottom', fontsize=8)
            
            # 图2: 网络数量对比
            bars2 = ax2.bar(x_pos, network_counts, color='#27ae60', alpha=0.7)
            ax2.set_xlabel('采集点位', fontsize=10)
            ax2.set_ylabel('网络数量', fontsize=10)
            ax2.set_title('检测网络数量对比', fontsize=12, fontweight='bold')
            ax2.set_xticks(x_pos)
            ax2.set_xticklabels(locations, rotation=15, ha='right', fontsize=9)
            ax2.grid(axis='y', alpha=0.3)
            
            for bar, val in zip(bars2, network_counts):
                ax2.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.5, 
                        f'{val}', ha='center', va='bottom', fontsize=9)
            
            # 图3: 强信号网络数量
            bars3 = ax3.bar(x_pos, strong_counts, color='#f39c12', alpha=0.7)
            ax3.set_xlabel('采集点位', fontsize=10)
            ax3.set_ylabel('强信号网络数 (>70%)', fontsize=10)
            ax3.set_title('强信号网络分布', fontsize=12, fontweight='bold')
            ax3.set_xticks(x_pos)
            ax3.set_xticklabels(locations, rotation=15, ha='right', fontsize=9)
            ax3.grid(axis='y', alpha=0.3)
            
            for bar, val in zip(bars3, strong_counts):
                ax3.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.3, 
                        f'{val}', ha='center', va='bottom', fontsize=9)
            
            # 图4: 综合评分雷达图
            from matplotlib.patches import Circle, RegularPolygon
            from matplotlib.path import Path
            from matplotlib.projections.polar import PolarAxes
            from matplotlib.projections import register_projection
            import numpy as np
            
            # 准备雷达图数据（归一化到0-100）
            categories = ['平均信号', '最强信号', '网络数量', '强信号占比']
            
            # 计算每个点位的综合指标
            ax4.set_title('综合评分雷达图', fontsize=12, fontweight='bold')
            angles = np.linspace(0, 2 * np.pi, len(categories), endpoint=False).tolist()
            angles += angles[:1]  # 闭合
            
            ax4_polar = plt.subplot(224, projection='polar')
            ax4_polar.set_theta_offset(np.pi / 2)
            ax4_polar.set_theta_direction(-1)
            ax4_polar.set_xticks(angles[:-1])
            ax4_polar.set_xticklabels(categories, fontsize=9)
            ax4_polar.set_ylim(0, 100)
            ax4_polar.set_yticks([25, 50, 75, 100])
            ax4_polar.grid(True)
            
            colors = plt.cm.Set3(np.linspace(0, 1, len(self.collection_points)))
            
            for idx, (point, color) in enumerate(zip(self.collection_points, colors)):
                if point['networks']:
                    signals = [net.get('signal_percent', 0) for net in point['networks']]
                    avg_sig = sum(signals) / len(signals)
                    max_sig = max(signals)
                    net_count_norm = min(len(signals) / 20 * 100, 100)  # 假设20个网络为满分
                    strong_ratio = sum(1 for s in signals if s > 70) / len(signals) * 100
                    
                    values = [avg_sig, max_sig, net_count_norm, strong_ratio]
                    values += values[:1]  # 闭合
                    
                    ax4_polar.plot(angles, values, 'o-', linewidth=2, label=point['location'], color=color)
                    ax4_polar.fill(angles, values, alpha=0.15, color=color)
            
            ax4_polar.legend(loc='upper right', bbox_to_anchor=(1.3, 1.0), fontsize=8)
            
            # 移除ax4（被polar替代）
            fig.delaxes(ax4)
            
            plt.tight_layout()
            
            # 保存图表
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            chart_file = f"多点位对比图表_{timestamp}.png"
            
            filepath = filedialog.asksaveasfilename(
                defaultextension=".png",
                filetypes=[("PNG图片", "*.png"), ("所有文件", "*.*")],
                initialfile=chart_file
            )
            
            if filepath:
                plt.savefig(filepath, dpi=300, bbox_inches='tight')
                self.logger.info(f"对比图表已保存到: {filepath}")
                messagebox.showinfo("成功", f"对比图表已保存到:\n{filepath}")
                
                # 显示图表
                plt.show()
            else:
                # 仅显示不保存
                plt.show()
            
        except ImportError as e:
            messagebox.showerror("错误", "图表生成功能需要matplotlib库\n请运行: pip install matplotlib")
        except Exception as e:
            self.logger.error(f"生成对比图表失败: {e}")
            messagebox.showerror("错误", f"生成图表失败: {str(e)}")
    
    def get_frame(self):
        """获取Frame对象"""
        return self.frame


